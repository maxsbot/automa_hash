import os
import hashlib
import openpyxl


# Função para calcular o hash de um arquivo usando SHA-256
def calcular_hash_arquivo(caminho_arquivo):
    sha256_hash = hashlib.sha256()
    with open(caminho_arquivo, "rb") as arquivo:
        for byte_em_chunk in iter(lambda: arquivo.read(4096), b""):
            sha256_hash.update(byte_em_chunk)
    return sha256_hash.hexdigest()


# Função para criar a tabela de hash a partir dos arquivos da pasta fornecida
def criar_tabela_hash(caminho_pasta):
    workbook = openpyxl.Workbook()  # Cria um novo arquivo .xlsx
    sheet = workbook.active  # Obtém a planilha ativa do arquivo
    # Define o cabeçalho das colunas
    sheet["A1"] = "Caminho do Arquivo"
    sheet["B1"] = "Nome do Arquivo"
    sheet["C1"] = "Tamanho (bytes)"
    sheet["D1"] = "Código de Autenticação (Hash)"

    row = 2  # Iniciar na linha 2 para deixar espaço para o cabeçalho
    total_arquivos = 0

    for pasta_raiz, sub_pastas, arquivos in os.walk(caminho_pasta):
        # Percorre os arquivos na árvore de diretórios da pasta fornecida
        for nome_arquivo in arquivos:
            if nome_arquivo == "Thumbs.db":
                continue  # Ignora o arquivo "Thumbs.db"

            caminho_arquivo = os.path.join(pasta_raiz, nome_arquivo)
            # Calcula o hash do arquivo usando a função calcular_hash_arquivo()
            hash_arquivo = calcular_hash_arquivo(caminho_arquivo)
            # Obtém o tamanho do arquivo em bytes
            tamanho_arquivo = os.path.getsize(caminho_arquivo)
            # Substitui o caminho da pasta pelo caractere "*" e o nome do arquivo por ""
            caminho_arquivo = caminho_arquivo.replace(caminho_pasta, "*").replace(
                nome_arquivo, ""
            )
            # Adiciona o caminho do arquivo, nome, tamanho e hash nas células correspondentes da planilha
            sheet.cell(row=row, column=1, value=caminho_arquivo)
            sheet.cell(row=row, column=2, value=nome_arquivo)
            sheet.cell(row=row, column=3, value=tamanho_arquivo)
            sheet.cell(row=row, column=4, value=hash_arquivo)
            row += 1
            total_arquivos += 1
            print(f"Arquivo {total_arquivos}: {nome_arquivo}")

    # Verifica se existem linhas com "Thumbs.db" na coluna "Nome do Arquivo" e as remove
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        if row[0].value == "Thumbs.db":
            sheet.delete_rows(row[0].row)

    nome_pasta = os.path.basename(caminho_pasta)
    nome_arquivo_xlsx = f"(Hash) {nome_pasta}.xlsx"
    workbook.save(nome_arquivo_xlsx)  # Salva o arquivo .xlsx com o nome formatado
    print(f"Tabela hash criada com sucesso no arquivo: {nome_arquivo_xlsx}")
    print(f"Total de arquivos processados: {total_arquivos}")


# Solicita ao usuário que digite o caminho da pasta
caminho_pasta = input("Digite o caminho da pasta: ")
# Chama a função criar_tabela_hash() passando o caminho da pasta fornecido
criar_tabela_hash(caminho_pasta)
